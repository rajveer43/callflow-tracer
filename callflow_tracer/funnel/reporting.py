"""
Funnel Export and Reporting System for CallFlow Tracer

This module provides comprehensive export and reporting capabilities for funnel analysis,
including multiple export formats, automated report generation, and scheduled reporting.
"""

import json
import csv
import io
import base64
from typing import Dict, List, Any, Optional, Union, BinaryIO
from dataclasses import asdict
from datetime import datetime, timedelta
from pathlib import Path
import tempfile
import zipfile

from .analysis import FunnelAnalyzer, FunnelStep, FunnelSession
from .visualizer import FunnelVisualizer
from .algorithms import FunnelAnomalyDetector, FunnelPatternRecognizer, FunnelOptimizer
from .monitor import RealTimeFunnelMonitor, FunnelDashboard


class FunnelExporter:
    """Advanced funnel data exporter"""

    def __init__(self, analyzer: FunnelAnalyzer):
        self.analyzer = analyzer
        self.supported_formats = ["json", "csv", "excel", "pdf", "html", "xml"]

    def export_data(
        self,
        format_type: str = "json",
        include_sections: Optional[List[str]] = None,
        output_path: Optional[str] = None,
    ) -> Union[str, bytes]:
        """Export funnel data in specified format"""
        if format_type not in self.supported_formats:
            raise ValueError(
                f"Unsupported format: {format_type}. Supported: {self.supported_formats}"
            )

        # Determine which sections to include
        if include_sections is None:
            include_sections = ["analytics", "steps", "sessions", "config"]

        # Gather data
        data = self._gather_export_data(include_sections)

        # Export based on format
        if format_type == "json":
            return self._export_json(data, output_path)
        elif format_type == "csv":
            return self._export_csv(data, output_path)
        elif format_type == "excel":
            return self._export_excel(data, output_path)
        elif format_type == "html":
            return self._export_html(data, output_path)
        elif format_type == "xml":
            return self._export_xml(data, output_path)
        elif format_type == "pdf":
            return self._export_pdf(data, output_path)

        raise ValueError(f"Export not implemented for format: {format_type}")

    def _gather_export_data(self, include_sections: List[str]) -> Dict[str, Any]:
        """Gather data for export based on sections"""
        data = {
            "export_info": {
                "timestamp": datetime.now().isoformat(),
                "funnel_name": self.analyzer.name,
                "funnel_type": self.analyzer.funnel_type.value,
                "funnel_id": self.analyzer.funnel_id,
                "total_steps": len(self.analyzer.steps),
                "total_sessions": len(self.analyzer.sessions),
                "active_sessions": len(self.analyzer.active_sessions),
            }
        }

        if "analytics" in include_sections:
            data["analytics"] = self.analyzer.get_analytics()

        if "steps" in include_sections:
            data["steps"] = [step.to_dict() for step in self.analyzer.steps]

        if "sessions" in include_sections:
            data["sessions"] = [
                session.to_dict() for session in self.analyzer.sessions.values()
            ]

        if "config" in include_sections:
            data["config"] = {
                "max_sessions": self.analyzer.max_sessions,
                "session_timeout_minutes": self.analyzer.session_timeout_minutes,
                "enable_real_time": self.analyzer.enable_real_time,
                "auto_detect_steps": self.analyzer.auto_detect_steps,
            }

        if "recommendations" in include_sections:
            from .algorithms import generate_optimization_plan

            data["recommendations"] = generate_optimization_plan(self.analyzer)

        if "anomalies" in include_sections:
            from .algorithms import analyze_funnel_anomalies

            detector = FunnelAnomalyDetector(self.analyzer)
            data["anomalies"] = [
                anomaly.__dict__ for anomaly in detector.detect_anomalies()
            ]

        if "patterns" in include_sections:
            from .algorithms import recognize_funnel_patterns

            recognizer = FunnelPatternRecognizer(self.analyzer)
            data["patterns"] = [
                pattern.__dict__ for pattern in recognizer.recognize_patterns()
            ]

        return data

    def _export_json(self, data: Dict[str, Any], output_path: Optional[str]) -> str:
        """Export data as JSON"""
        json_data = json.dumps(data, indent=2, default=str)

        if output_path:
            with open(output_path, "w") as f:
                f.write(json_data)
            return output_path

        return json_data

    def _export_csv(self, data: Dict[str, Any], output_path: Optional[str]) -> str:
        """Export data as CSV"""
        output = io.StringIO()

        # Export steps data
        if "steps" in data:
            writer = csv.writer(output)
            writer.writerow(
                [
                    "Step Name",
                    "Order",
                    "Total Users",
                    "Successful",
                    "Failed",
                    "Conversion Rate",
                    "Dropoff Rate",
                    "Avg Time (ms)",
                    "Error Rate",
                ]
            )

            for step in data["steps"]:
                writer.writerow(
                    [
                        step.get("name", ""),
                        step.get("order", 0),
                        step.get("total_users", 0),
                        step.get("successful_users", 0),
                        step.get("failed_users", 0),
                        step.get("conversion_rate", 0),
                        step.get("dropoff_rate", 0),
                        step.get("avg_time_ms", 0),
                        step.get("error_rate", 0),
                    ]
                )

        csv_data = output.getvalue()

        if output_path:
            with open(output_path, "w") as f:
                f.write(csv_data)
            return output_path

        return csv_data

    def _export_excel(self, data: Dict[str, Any], output_path: Optional[str]) -> bytes:
        """Export data as Excel file"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.chart import BarChart, Reference

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create sheets
            if "steps" in data:
                ws_steps = wb.create_sheet("Steps")

                # Headers
                headers = [
                    "Step Name",
                    "Order",
                    "Total Users",
                    "Successful",
                    "Failed",
                    "Conversion Rate",
                    "Dropoff Rate",
                    "Avg Time (ms)",
                    "Error Rate",
                ]
                ws_steps.append(headers)

                # Style headers
                for col in range(1, len(headers) + 1):
                    cell = ws_steps.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
                    )

                # Data
                for step in data["steps"]:
                    ws_steps.append(
                        [
                            step.get("name", ""),
                            step.get("order", 0),
                            step.get("total_users", 0),
                            step.get("successful_users", 0),
                            step.get("failed_users", 0),
                            step.get("conversion_rate", 0),
                            step.get("dropoff_rate", 0),
                            step.get("avg_time_ms", 0),
                            step.get("error_rate", 0),
                        ]
                    )

                # Add chart
                if len(data["steps"]) > 1:
                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.title = "Conversion Rates by Step"
                    chart.y_axis.title = "Conversion Rate (%)"
                    chart.x_axis.title = "Steps"

                    data_ref = Reference(
                        ws_steps,
                        min_col=6,
                        min_row=1,
                        max_col=6,
                        max_row=len(data["steps"]) + 1,
                    )
                    cats_ref = Reference(
                        ws_steps, min_col=1, min_row=2, max_row=len(data["steps"]) + 1
                    )

                    chart.add_data(data_ref, titles_from_data=True)
                    chart.set_categories(cats_ref)
                    chart.height = 10
                    chart.width = 15

                    ws_steps.add_chart(chart, "K2")

            if "sessions" in data:
                ws_sessions = wb.create_sheet("Sessions")

                # Headers
                headers = [
                    "Session ID",
                    "User ID",
                    "Start Time",
                    "End Time",
                    "Status",
                    "Completed Steps",
                    "Failed Steps",
                    "Total Duration (ms)",
                ]
                ws_sessions.append(headers)

                # Style headers
                for col in range(1, len(headers) + 1):
                    cell = ws_sessions.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
                    )

                # Data
                for session in data["sessions"]:
                    ws_sessions.append(
                        [
                            session.get("session_id", ""),
                            session.get("user_id", ""),
                            session.get("start_time", ""),
                            session.get("end_time", ""),
                            session.get("status", ""),
                            len(session.get("completed_steps", [])),
                            len(session.get("failed_steps", [])),
                            session.get("total_duration_ms", 0),
                        ]
                    )

            if "analytics" in data:
                ws_analytics = wb.create_sheet("Analytics")

                # Key metrics
                analytics = data["analytics"]
                metrics = [
                    (
                        "Overall Conversion Rate",
                        analytics.get("conversion_metrics", {}).get(
                            "overall_conversion_rate", 0
                        ),
                    ),
                    (
                        "Total Sessions",
                        analytics.get("conversion_metrics", {}).get(
                            "total_sessions", 0
                        ),
                    ),
                    (
                        "Completed Sessions",
                        analytics.get("conversion_metrics", {}).get(
                            "completed_sessions", 0
                        ),
                    ),
                    (
                        "Average Step Time",
                        analytics.get("performance_metrics", {}).get(
                            "average_step_time_ms", 0
                        ),
                    ),
                    (
                        "Total Errors",
                        analytics.get("error_analysis", {}).get("total_errors", 0),
                    ),
                    (
                        "Error Rate",
                        analytics.get("error_analysis", {}).get("error_rate", 0),
                    ),
                ]

                ws_analytics.append(["Metric", "Value"])
                for metric, value in metrics:
                    ws_analytics.append([metric, value])

            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_data = excel_buffer.getvalue()

            if output_path:
                with open(output_path, "wb") as f:
                    f.write(excel_data)
                return output_path

            return excel_data

        except ImportError:
            # Fallback to CSV if pandas/openpyxl not available
            return self._export_csv(data, output_path).encode("utf-8")

    def _export_html(self, data: Dict[str, Any], output_path: Optional[str]) -> str:
        """Export data as HTML report"""
        html_template = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Funnel Analysis Report - {funnel_name}</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background-color: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ text-align: center; margin-bottom: 40px; border-bottom: 2px solid #3b82f6; padding-bottom: 20px; }}
        .header h1 {{ color: #1f2937; margin: 0; }}
        .header .subtitle {{ color: #6b7280; margin: 10px 0 0 0; }}
        .section {{ margin: 30px 0; }}
        .section h2 {{ color: #1f2937; border-bottom: 1px solid #e5e7eb; padding-bottom: 10px; }}
        .metrics {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }}
        .metric {{ background: #f8fafc; padding: 20px; border-radius: 8px; text-align: center; }}
        .metric .value {{ font-size: 2em; font-weight: bold; color: #3b82f6; }}
        .metric .label {{ color: #6b7280; margin-top: 5px; }}
        .table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        .table th, .table td {{ padding: 12px; text-align: left; border-bottom: 1px solid #e5e7eb; }}
        .table th {{ background: #f8fafc; font-weight: 600; color: #1f2937; }}
        .table tr:hover {{ background: #f8fafc; }}
        .chart {{ margin: 20px 0; text-align: center; }}
        .funnel-bar {{ background: linear-gradient(to right, #3b82f6, #60a5fa); height: 40px; margin: 10px 0; border-radius: 4px; display: flex; align-items: center; padding: 0 15px; color: white; font-weight: bold; }}
        .alert {{ padding: 15px; border-radius: 4px; margin: 10px 0; }}
        .alert.critical {{ background: #fef2f2; border: 1px solid #fecaca; color: #991b1b; }}
        .alert.warning {{ background: #fffbeb; border: 1px solid #fed7aa; color: #92400e; }}
        .alert.info {{ background: #eff6ff; border: 1px solid #bfdbfe; color: #1e40af; }}
        footer {{ text-align: center; margin-top: 40px; padding-top: 20px; border-top: 1px solid #e5e7eb; color: #6b7280; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Funnel Analysis Report</h1>
            <div class="subtitle">{funnel_name} - Generated on {timestamp}</div>
        </div>
        
        {content}
        
        <footer>
            <p>Generated by CallFlow Tracer - Funnel Analysis System</p>
        </footer>
    </div>
</body>
</html>
        """

        # Build content sections
        content_sections = []

        # Key metrics section
        if "analytics" in data:
            analytics = data["analytics"]
            conversion_metrics = analytics.get("conversion_metrics", {})
            performance_metrics = analytics.get("performance_metrics", {})
            error_analysis = analytics.get("error_analysis", {})

            metrics_html = """
        <div class="section">
            <h2>Key Metrics</h2>
            <div class="metrics">
                <div class="metric">
                    <div class="value">{conversion_rate:.1f}%</div>
                    <div class="label">Overall Conversion Rate</div>
                </div>
                <div class="metric">
                    <div class="value">{total_sessions}</div>
                    <div class="label">Total Sessions</div>
                </div>
                <div class="metric">
                    <div class="value">{avg_time:.0f}ms</div>
                    <div class="label">Average Step Time</div>
                </div>
                <div class="metric">
                    <div class="value">{error_rate:.1f}%</div>
                    <div class="label">Error Rate</div>
                </div>
            </div>
        </div>
            """.format(
                conversion_rate=conversion_metrics.get("overall_conversion_rate", 0),
                total_sessions=conversion_metrics.get("total_sessions", 0),
                avg_time=performance_metrics.get("average_step_time_ms", 0),
                error_rate=error_analysis.get("error_rate", 0),
            )
            content_sections.append(metrics_html)

        # Funnel visualization
        if "steps" in data:
            steps = data["steps"]
            max_users = (
                max(step.get("total_users", 0) for step in steps) if steps else 1
            )

            funnel_html = """
        <div class="section">
            <h2>Funnel Visualization</h2>
            <div class="chart">
                {funnel_bars}
            </div>
        </div>
            """.format(
                funnel_bars="\n".join(
                    [
                        f'<div class="funnel-bar" style="width: {(step.get("total_users", 0) / max_users) * 100}%">'
                        f'{step.get("name", "")}: {step.get("total_users", 0)} users ({step.get("conversion_rate", 0):.1f}% conversion)'
                        f"</div>"
                        for step in steps
                    ]
                )
            )
            content_sections.append(funnel_html)

        # Steps table
        if "steps" in data:
            steps_html = """
        <div class="section">
            <h2>Step Details</h2>
            <table class="table">
                <thead>
                    <tr>
                        <th>Step Name</th>
                        <th>Total Users</th>
                        <th>Successful</th>
                        <th>Failed</th>
                        <th>Conversion Rate</th>
                        <th>Avg Time (ms)</th>
                        <th>Error Rate</th>
                    </tr>
                </thead>
                <tbody>
                    {step_rows}
                </tbody>
            </table>
        </div>
            """.format(
                step_rows="\n".join(
                    [
                        f"<tr>"
                        f'<td>{step.get("name", "")}</td>'
                        f'<td>{step.get("total_users", 0)}</td>'
                        f'<td>{step.get("successful_users", 0)}</td>'
                        f'<td>{step.get("failed_users", 0)}</td>'
                        f'<td>{step.get("conversion_rate", 0):.1f}%</td>'
                        f'<td>{step.get("avg_time_ms", 0):.0f}</td>'
                        f'<td>{step.get("error_rate", 0):.1f}%</td>'
                        f"</tr>"
                        for step in steps
                    ]
                )
            )
            content_sections.append(steps_html)

        # Recommendations
        if "recommendations" in data:
            recommendations = data["recommendations"].get("recommendations", [])
            if recommendations:
                rec_html = """
        <div class="section">
            <h2>Optimization Recommendations</h2>
            {recommendation_alerts}
        </div>
                """.format(
                    recommendation_alerts="\n".join(
                        [
                            f'<div class="alert {rec.get("priority", "info")}">'
                            f'<strong>{rec.get("title", "")}</strong><br>'
                            f'{rec.get("description", "")}'
                            f"</div>"
                            for rec in recommendations[:5]
                        ]
                    )
                )
                content_sections.append(rec_html)

        # Generate final HTML
        html_content = html_template.format(
            funnel_name=data["export_info"]["funnel_name"],
            timestamp=data["export_info"]["timestamp"],
            content="\n".join(content_sections),
        )

        if output_path:
            with open(output_path, "w") as f:
                f.write(html_content)
            return output_path

        return html_content

    def _export_xml(self, data: Dict[str, Any], output_path: Optional[str]) -> str:
        """Export data as XML"""

        def dict_to_xml(d, root_name="root"):
            """Convert dictionary to XML"""
            xml = f"<{root_name}>"
            for key, value in d.items():
                if isinstance(value, dict):
                    xml += dict_to_xml(value, key)
                elif isinstance(value, list):
                    for item in value:
                        if isinstance(item, dict):
                            xml += dict_to_xml(
                                item, key.rstrip("s")
                            )  # Remove plural 's'
                        else:
                            xml += f"<{key}>{str(item)}</{key}>"
                else:
                    xml += f"<{key}>{str(value)}</{key}>"
            xml += f"</{root_name}>"
            return xml

        xml_data = dict_to_xml(data, "funnel_export")

        if output_path:
            with open(output_path, "w") as f:
                f.write(xml_data)
            return output_path

        return xml_data

    def _export_pdf(self, data: Dict[str, Any], output_path: Optional[str]) -> bytes:
        """Export data as PDF"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import (
                SimpleDocTemplate,
                Paragraph,
                Spacer,
                Table,
                TableStyle,
                PageBreak,
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.barcharts import VerticalBarChart

            # Create PDF buffer
            pdf_buffer = io.BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)

            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=24,
                spaceAfter=30,
                alignment=1,  # Center
            )

            # Build content
            story = []

            # Title
            story.append(Paragraph(f"Funnel Analysis Report", title_style))
            story.append(
                Paragraph(f"{data['export_info']['funnel_name']}", styles["Heading2"])
            )
            story.append(
                Paragraph(
                    f"Generated: {data['export_info']['timestamp']}", styles["Normal"]
                )
            )
            story.append(Spacer(1, 20))

            # Key metrics
            if "analytics" in data:
                analytics = data["analytics"]
                conversion_metrics = analytics.get("conversion_metrics", {})
                performance_metrics = analytics.get("performance_metrics", {})
                error_analysis = analytics.get("error_analysis", {})

                story.append(Paragraph("Key Metrics", styles["Heading2"]))

                metrics_data = [
                    ["Metric", "Value"],
                    [
                        "Overall Conversion Rate",
                        f"{conversion_metrics.get('overall_conversion_rate', 0):.1f}%",
                    ],
                    [
                        "Total Sessions",
                        str(conversion_metrics.get("total_sessions", 0)),
                    ],
                    [
                        "Completed Sessions",
                        str(conversion_metrics.get("completed_sessions", 0)),
                    ],
                    [
                        "Average Step Time",
                        f"{performance_metrics.get('average_step_time_ms', 0):.0f}ms",
                    ],
                    ["Total Errors", str(error_analysis.get("total_errors", 0))],
                    ["Error Rate", f"{error_analysis.get('error_rate', 0):.1f}%"],
                ]

                metrics_table = Table(metrics_data)
                metrics_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 14),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]
                    )
                )

                story.append(metrics_table)
                story.append(Spacer(1, 20))

            # Steps table
            if "steps" in data:
                story.append(Paragraph("Step Details", styles["Heading2"]))

                steps_data = [
                    [
                        "Step Name",
                        "Total Users",
                        "Successful",
                        "Failed",
                        "Conversion Rate",
                        "Avg Time (ms)",
                        "Error Rate",
                    ]
                ]

                for step in data["steps"]:
                    steps_data.append(
                        [
                            step.get("name", ""),
                            str(step.get("total_users", 0)),
                            str(step.get("successful_users", 0)),
                            str(step.get("failed_users", 0)),
                            f"{step.get('conversion_rate', 0):.1f}%",
                            f"{step.get('avg_time_ms', 0):.0f}",
                            f"{step.get('error_rate', 0):.1f}%",
                        ]
                    )

                steps_table = Table(steps_data)
                steps_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("FONTSIZE", (0, 0), (-1, 0), 12),
                            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]
                    )
                )

                story.append(steps_table)

            # Build PDF
            doc.build(story)
            pdf_data = pdf_buffer.getvalue()

            if output_path:
                with open(output_path, "wb") as f:
                    f.write(pdf_data)
                return output_path

            return pdf_data

        except ImportError:
            # Fallback to HTML if reportlab not available
            html_data = self._export_html(data, None)
            return html_data.encode("utf-8")


class FunnelReporter:
    """Advanced funnel reporting system"""

    def __init__(self, analyzer: FunnelAnalyzer):
        self.analyzer = analyzer
        self.exporter = FunnelExporter(analyzer)
        self.visualizer = FunnelVisualizer(analyzer)

    def generate_comprehensive_report(
        self,
        report_type: str = "standard",
        include_sections: Optional[List[str]] = None,
        output_path: Optional[str] = None,
    ) -> str:
        """Generate comprehensive funnel report"""
        if include_sections is None:
            include_sections = [
                "executive_summary",
                "analytics",
                "steps",
                "sessions",
                "anomalies",
                "patterns",
                "recommendations",
                "appendix",
            ]

        # Gather all data
        report_data = self._gather_report_data(include_sections)

        # Generate report based on type
        if report_type == "standard":
            return self._generate_standard_report(report_data, output_path)
        elif report_type == "executive":
            return self._generate_executive_report(report_data, output_path)
        elif report_type == "technical":
            return self._generate_technical_report(report_data, output_path)
        elif report_type == "comparison":
            return self._generate_comparison_report(report_data, output_path)
        else:
            raise ValueError(f"Unsupported report type: {report_type}")

    def _gather_report_data(self, include_sections: List[str]) -> Dict[str, Any]:
        """Gather comprehensive data for reporting"""
        data = {
            "report_info": {
                "generated_at": datetime.now().isoformat(),
                "funnel_name": self.analyzer.name,
                "funnel_type": self.analyzer.funnel_type.value,
                "report_sections": include_sections,
            }
        }

        # Executive summary
        if "executive_summary" in include_sections:
            data["executive_summary"] = self._generate_executive_summary()

        # Analytics
        if "analytics" in include_sections:
            data["analytics"] = self.analyzer.get_analytics()

        # Detailed steps
        if "steps" in include_sections:
            data["steps"] = [step.to_dict() for step in self.analyzer.steps]

        # Sessions
        if "sessions" in include_sections:
            data["sessions"] = [
                session.to_dict() for session in self.analyzer.sessions.values()
            ]

        # Anomalies
        if "anomalies" in include_sections:
            from .algorithms import analyze_funnel_anomalies

            detector = FunnelAnomalyDetector(self.analyzer)
            data["anomalies"] = [
                anomaly.__dict__ for anomaly in detector.detect_anomalies()
            ]

        # Patterns
        if "patterns" in include_sections:
            from .algorithms import recognize_funnel_patterns

            recognizer = FunnelPatternRecognizer(self.analyzer)
            data["patterns"] = [
                pattern.__dict__ for pattern in recognizer.recognize_patterns()
            ]

        # Recommendations
        if "recommendations" in include_sections:
            from .algorithms import generate_optimization_plan

            data["recommendations"] = generate_optimization_plan(self.analyzer)

        # Visualizations
        if "visualizations" in include_sections:
            data["visualizations"] = {
                "funnel_chart": self.visualizer.generate_funnel_chart(),
                "performance_chart": self.visualizer.generate_performance_chart(),
                "error_chart": self.visualizer.generate_error_analysis_chart(),
                "dashboard": self.visualizer.generate_dashboard(),
            }

        # Appendix
        if "appendix" in include_sections:
            data["appendix"] = {
                "methodology": self._generate_methodology_section(),
                "data_quality": self._assess_data_quality(),
                "limitations": self._identify_limitations(),
            }

        return data

    def _generate_executive_summary(self) -> Dict[str, Any]:
        """Generate executive summary"""
        analytics = self.analyzer.get_analytics()
        conversion_metrics = analytics.get("conversion_metrics", {})
        performance_metrics = analytics.get("performance_metrics", {})

        return {
            "overview": {
                "total_sessions": conversion_metrics.get("total_sessions", 0),
                "conversion_rate": conversion_metrics.get("overall_conversion_rate", 0),
                "average_step_time": performance_metrics.get("average_step_time_ms", 0),
                "total_errors": analytics.get("error_analysis", {}).get(
                    "total_errors", 0
                ),
            },
            "key_findings": self._extract_key_findings(analytics),
            "business_impact": self._assess_business_impact(analytics),
            "next_steps": self._recommend_next_steps(analytics),
        }

    def _extract_key_findings(self, analytics: Dict[str, Any]) -> List[str]:
        """Extract key findings from analytics"""
        findings = []

        conversion_metrics = analytics.get("conversion_metrics", {})
        overall_rate = conversion_metrics.get("overall_conversion_rate", 0)

        if overall_rate < 20:
            findings.append(
                f"Low overall conversion rate ({overall_rate:.1f}%) requires immediate attention"
            )
        elif overall_rate > 80:
            findings.append(
                f"Excellent conversion rate ({overall_rate:.1f}%) indicates good user experience"
            )

        # Check for biggest dropoff
        biggest_dropoff = conversion_metrics.get("biggest_dropoff")
        if biggest_dropoff and biggest_dropoff.get("relative_dropoff", 0) > 30:
            findings.append(
                f"Significant dropoff at step '{biggest_dropoff.get('step_name')}' ({biggest_dropoff.get('relative_dropoff', 0):.1f}%)"
            )

        # Performance issues
        performance_metrics = analytics.get("performance_metrics", {})
        avg_time = performance_metrics.get("average_step_time_ms", 0)
        if avg_time > 5000:
            findings.append(
                f"Performance issues detected with average step time of {avg_time:.0f}ms"
            )

        # Error rates
        error_analysis = analytics.get("error_analysis", {})
        error_rate = error_analysis.get("error_rate", 0)
        if error_rate > 10:
            findings.append(
                f"High error rate ({error_rate:.1f}%) affecting user experience"
            )

        return findings

    def _assess_business_impact(self, analytics: Dict[str, Any]) -> Dict[str, Any]:
        """Assess business impact"""
        conversion_metrics = analytics.get("conversion_metrics", {})
        total_sessions = conversion_metrics.get("total_sessions", 0)
        completed_sessions = conversion_metrics.get("completed_sessions", 0)

        # Estimate revenue impact (simplified)
        lost_opportunities = total_sessions - completed_sessions
        estimated_revenue_loss = lost_opportunities * 100  # Assume $100 per conversion

        return {
            "revenue_impact": {
                "lost_opportunities": lost_opportunities,
                "estimated_revenue_loss": estimated_revenue_loss,
                "potential_improvement_value": completed_sessions
                * 0.2
                * 100,  # 20% improvement potential
            },
            "operational_impact": {
                "support_tickets_reduction": analytics.get("error_analysis", {}).get(
                    "total_errors", 0
                )
                * 0.5,
                "infrastructure_savings": analytics.get("performance_metrics", {}).get(
                    "average_step_time_ms", 0
                )
                * 0.01,
            },
            "customer_satisfaction": {
                "current_score": max(
                    0,
                    100 - analytics.get("error_analysis", {}).get("error_rate", 0) * 2,
                ),
                "potential_improvement": 15,
            },
        }

    def _recommend_next_steps(self, analytics: Dict[str, Any]) -> List[Dict[str, str]]:
        """Recommend next steps"""
        steps = []

        # Based on conversion rate
        conversion_rate = analytics.get("conversion_metrics", {}).get(
            "overall_conversion_rate", 0
        )
        if conversion_rate < 30:
            steps.append(
                {
                    "priority": "High",
                    "action": "Immediate conversion optimization",
                    "timeline": "1-2 weeks",
                    "owner": "Product Team",
                }
            )

        # Based on performance
        avg_time = analytics.get("performance_metrics", {}).get(
            "average_step_time_ms", 0
        )
        if avg_time > 3000:
            steps.append(
                {
                    "priority": "Medium",
                    "action": "Performance optimization",
                    "timeline": "2-4 weeks",
                    "owner": "Engineering Team",
                }
            )

        # Based on errors
        error_rate = analytics.get("error_analysis", {}).get("error_rate", 0)
        if error_rate > 5:
            steps.append(
                {
                    "priority": "High",
                    "action": "Error handling improvement",
                    "timeline": "1 week",
                    "owner": "Development Team",
                }
            )

        return steps

    def _generate_standard_report(
        self, data: Dict[str, Any], output_path: Optional[str]
    ) -> str:
        """Generate standard comprehensive report"""
        html_template = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Funnel Analysis Report - {funnel_name}</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; line-height: 1.6; margin: 0; padding: 20px; background: #f8fafc; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 40px; border-radius: 8px; box-shadow: 0 4px 20px rgba(0,0,0,0.1); }}
        .header {{ text-align: center; margin-bottom: 50px; border-bottom: 3px solid #3b82f6; padding-bottom: 30px; }}
        .header h1 {{ color: #1e293b; margin: 0; font-size: 2.5em; }}
        .header .subtitle {{ color: #64748b; margin: 10px 0 0 0; font-size: 1.1em; }}
        .toc {{ background: #f1f5f9; padding: 30px; border-radius: 8px; margin: 30px 0; }}
        .toc h3 {{ color: #1e293b; margin-top: 0; }}
        .toc ul {{ list-style: none; padding: 0; }}
        .toc li {{ margin: 8px 0; }}
        .toc a {{ color: #3b82f6; text-decoration: none; }}
        .toc a:hover {{ text-decoration: underline; }}
        .section {{ margin: 40px 0; }}
        .section h2 {{ color: #1e293b; border-bottom: 2px solid #e2e8f0; padding-bottom: 10px; margin-bottom: 20px; }}
        .section h3 {{ color: #334155; margin-top: 30px; }}
        .highlight-box {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 8px; margin: 20px 0; }}
        .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; margin: 30px 0; }}
        .metric-card {{ background: #f8fafc; padding: 25px; border-radius: 8px; text-align: center; border-left: 4px solid #3b82f6; }}
        .metric-value {{ font-size: 2.5em; font-weight: bold; color: #3b82f6; margin-bottom: 10px; }}
        .metric-label {{ color: #64748b; font-weight: 500; }}
        .table {{ width: 100%; border-collapse: collapse; margin: 20px 0; background: white; }}
        .table th, .table td {{ padding: 15px; text-align: left; border-bottom: 1px solid #e2e8f0; }}
        .table th {{ background: #f8fafc; font-weight: 600; color: #1e293b; }}
        .table tr:hover {{ background: #f8fafc; }}
        .recommendation {{ background: #eff6ff; border: 1px solid #bfdbfe; padding: 20px; border-radius: 8px; margin: 15px 0; }}
        .recommendation h4 {{ color: #1e40af; margin-top: 0; }}
        .finding {{ background: #fef3c7; border: 1px solid #fcd34d; padding: 20px; border-radius: 8px; margin: 15px 0; }}
        .finding h4 {{ color: #92400e; margin-top: 0; }}
        .chart-placeholder {{ background: #f1f5f9; padding: 40px; text-align: center; border-radius: 8px; margin: 20px 0; color: #64748b; }}
        footer {{ text-align: center; margin-top: 50px; padding-top: 30px; border-top: 1px solid #e2e8f0; color: #64748b; }}
        @media print {{ body {{ background: white; }} .container {{ box-shadow: none; }} }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Funnel Analysis Report</h1>
            <div class="subtitle">{funnel_name} • Generated on {timestamp}</div>
        </div>
        
        {content}
        
        <footer>
            <p><strong>CallFlow Tracer - Funnel Analysis System</strong><br>
            Advanced Performance & Conversion Analytics</p>
        </footer>
    </div>
</body>
</html>
        """

        # Build content sections
        content_sections = []

        # Table of Contents
        toc_html = """
        <div class="toc">
            <h3>Table of Contents</h3>
            <ul>
                <li><a href="#executive-summary">Executive Summary</a></li>
                <li><a href="#key-metrics">Key Metrics</a></li>
                <li><a href="#funnel-analysis">Funnel Analysis</a></li>
                <li><a href="#performance-analysis">Performance Analysis</a></li>
                <li><a href="#anomalies">Anomalies & Issues</a></li>
                <li><a href="#recommendations">Recommendations</a></li>
                <li><a href="#appendix">Appendix</a></li>
            </ul>
        </div>
        """
        content_sections.append(toc_html)

        # Executive Summary
        if "executive_summary" in data:
            exec_summary = data["executive_summary"]
            exec_html = f"""
        <div id="executive-summary" class="section">
            <h2>Executive Summary</h2>
            <div class="highlight-box">
                <h3>Overview</h3>
                <div class="metric-grid">
                    <div class="metric-card">
                        <div class="metric-value">{exec_summary['overview']['conversion_rate']:.1f}%</div>
                        <div class="metric-label">Conversion Rate</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">{exec_summary['overview']['total_sessions']:,}</div>
                        <div class="metric-label">Total Sessions</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">{exec_summary['overview']['average_step_time']:.0f}ms</div>
                        <div class="metric-label">Avg Step Time</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">{exec_summary['overview']['total_errors']}</div>
                        <div class="metric-label">Total Errors</div>
                    </div>
                </div>
            </div>
            
            <h3>Key Findings</h3>
            {self._format_findings(exec_summary['key_findings'])}
            
            <h3>Business Impact</h3>
            {self._format_business_impact(exec_summary['business_impact'])}
            
            <h3>Next Steps</h3>
            {self._format_next_steps(exec_summary['next_steps'])}
        </div>
            """
            content_sections.append(exec_html)

        # Generate final HTML
        html_content = html_template.format(
            funnel_name=data["report_info"]["funnel_name"],
            timestamp=data["report_info"]["generated_at"],
            content="\n".join(content_sections),
        )

        if output_path:
            with open(output_path, "w") as f:
                f.write(html_content)
            return output_path

        return html_content

    def _format_findings(self, findings: List[str]) -> str:
        """Format findings as HTML"""
        return "\n".join(
            [
                f'<div class="finding"><h4>Finding</h4><p>{finding}</p></div>'
                for finding in findings
            ]
        )

    def _format_business_impact(self, impact: Dict[str, Any]) -> str:
        """Format business impact as HTML"""
        revenue_impact = impact.get("revenue_impact", {})
        return f"""
        <div class="metric-grid">
            <div class="metric-card">
                <div class="metric-value">${revenue_impact.get('estimated_revenue_loss', 0):,.0f}</div>
                <div class="metric-label">Estimated Revenue Loss</div>
            </div>
            <div class="metric-card">
                <div class="metric-value">${revenue_impact.get('potential_improvement_value', 0):,.0f}</div>
                <div class="metric-label">Potential Improvement Value</div>
            </div>
        </div>
        """

    def _format_next_steps(self, steps: List[Dict[str, str]]) -> str:
        """Format next steps as HTML"""
        return "\n".join(
            [
                f'<div class="recommendation">'
                f'<h4>{step["action"]} (Priority: {step["priority"]})</h4>'
                f'<p><strong>Timeline:</strong> {step["timeline"]} | <strong>Owner:</strong> {step["owner"]}</p>'
                f"</div>"
                for step in steps
            ]
        )

    def _generate_executive_report(
        self, data: Dict[str, Any], output_path: Optional[str]
    ) -> str:
        """Generate executive-focused report"""
        # Focus on business impact and high-level insights
        exec_data = {
            "report_info": data["report_info"],
            "executive_summary": data.get("executive_summary", {}),
            "recommendations": data.get("recommendations", {}),
            "business_impact": data.get("executive_summary", {}).get(
                "business_impact", {}
            ),
        }

        return self._generate_standard_report(exec_data, output_path)

    def _generate_technical_report(
        self, data: Dict[str, Any], output_path: Optional[str]
    ) -> str:
        """Generate technical-focused report"""
        # Focus on technical details and implementation
        tech_data = {
            "report_info": data["report_info"],
            "analytics": data.get("analytics", {}),
            "steps": data.get("steps", []),
            "sessions": data.get("sessions", []),
            "anomalies": data.get("anomalies", []),
            "patterns": data.get("patterns", []),
            "appendix": data.get("appendix", {}),
        }

        return self._generate_standard_report(tech_data, output_path)

    def _generate_comparison_report(
        self, data: Dict[str, Any], output_path: Optional[str]
    ) -> str:
        """Generate comparison report"""
        # Would need comparison data - placeholder implementation
        return self._generate_standard_report(data, output_path)

    def _generate_methodology_section(self) -> Dict[str, Any]:
        """Generate methodology section"""
        return {
            "data_collection": "Real-time funnel tracking with CallFlow Tracer",
            "analysis_methods": [
                "Statistical analysis of conversion rates",
                "Performance bottleneck identification",
                "Anomaly detection using statistical outliers",
                "Pattern recognition using machine learning algorithms",
                "Predictive analytics using time series analysis",
            ],
            "confidence_intervals": "95% confidence intervals for all statistical measures",
            "data_quality_checks": "Automated validation and outlier removal",
        }

    def _assess_data_quality(self) -> Dict[str, Any]:
        """Assess data quality"""
        total_sessions = len(self.analyzer.sessions)
        completed_sessions = sum(
            1 for s in self.analyzer.sessions.values() if s.is_completed
        )

        return {
            "completeness": f"{(completed_sessions / total_sessions * 100) if total_sessions > 0 else 0:.1f}%",
            "accuracy": "High - automated data collection",
            "consistency": "Consistent tracking across all steps",
            "timeliness": "Real-time data collection",
            "sample_size": total_sessions,
            "data_freshness": "Live data",
        }

    def _identify_limitations(self) -> List[str]:
        """Identify analysis limitations"""
        return [
            "Analysis based on available tracked data",
            "External factors (market conditions, seasonality) not considered",
            "User demographics and behavior patterns limited to tracked metrics",
            "Predictive accuracy depends on historical data patterns",
            "Recommendations based on statistical correlations, not causation",
        ]

    def schedule_report(
        self,
        report_type: str = "standard",
        schedule: str = "weekly",
        output_path: str = "funnel_report.html",
        email_recipients: Optional[List[str]] = None,
    ):
        """Schedule automated report generation"""
        # This would integrate with a scheduling system
        # Placeholder implementation
        schedule_info = {
            "report_type": report_type,
            "schedule": schedule,
            "output_path": output_path,
            "email_recipients": email_recipients or [],
            "next_run": self._calculate_next_run(schedule),
        }

        return schedule_info

    def _calculate_next_run(self, schedule: str) -> datetime:
        """Calculate next run time for scheduled report"""
        now = datetime.now()

        if schedule == "daily":
            return now + timedelta(days=1)
        elif schedule == "weekly":
            return now + timedelta(weeks=1)
        elif schedule == "monthly":
            return now + timedelta(days=30)
        else:
            return now + timedelta(hours=1)


# Convenience functions
def export_funnel_data(
    analyzer: FunnelAnalyzer,
    format_type: str = "json",
    output_path: Optional[str] = None,
) -> Union[str, bytes]:
    """Export funnel data"""
    exporter = FunnelExporter(analyzer)
    return exporter.export_data(format_type, output_path=output_path)


def generate_funnel_report(
    analyzer: FunnelAnalyzer,
    report_type: str = "standard",
    output_path: Optional[str] = None,
) -> str:
    """Generate funnel report"""
    reporter = FunnelReporter(analyzer)
    return reporter.generate_comprehensive_report(report_type, output_path=output_path)
